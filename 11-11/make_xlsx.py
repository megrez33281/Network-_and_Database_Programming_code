from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 

def create_excel_file(filename):
    # 建立 xlsx 文件
    workbook = Workbook()  # 建立新的 Workbook
    workbook.save(filename)  # 儲存為指定的檔案名稱

    

def write_to_excel(filename, data, sheet_name = "Sheet"):
    # 寫入xlxs檔案
    # 檢查檔案是否存在
    try:
        workbook = load_workbook(filename)  # 如果檔案存在，則載入
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]  # 使用指定的工作表
        else:
            worksheet = workbook.create_sheet(title=sheet_name)  # 新建指定名稱的工作表
    except FileNotFoundError:
        create_excel_file(filename)
        write_to_excel(filename, data, sheet_name)
        return
    
    # 寫入資料到工作表
    for row_index, row_data in enumerate(data, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=col_index, value=cell_value)

    
    # 儲存檔案
    workbook.save(filename)

